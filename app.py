#! Python 3
# - Copy and Paste Ranges using OpenPyXl library

import openpyxl

# Prepare the spreadsheets to copy from and paste too.

# File to be copied
wb = openpyxl.load_workbook("report.xlsx")  # Add file name
sheet = wb["Sheet1"]  # Add Sheet name

# File to be pasted into
template = openpyxl.load_workbook("importInvoice.xlsx")  # Add file name
temp_sheet = template["Sheet1"]  # Add Sheet name

print("123456++++>>>>.....")
# Copy range of cells as a nested list
# Takes: start cell, end cell, and sheet you want to copy from.
def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    # Loops through selected Rows
    for i in range(startRow, endRow + 1, 1):
        # Appends the row to a RowSelected list
        rowSelected = []
        for j in range(startCol, endCol + 1, 1):
            rowSelected.append(sheet.cell(row=i, column=j).value)
        # Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)

    return rangeSelected

print("I'm here....")

# Paste range
# Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow, endRow + 1, 1):
        countCol = 0
        for j in range(startCol, endCol + 1, 1):
            sheetReceiving.cell(row=i, column=j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1

print("I'm almost there...")

def copyOrderID():
    print("processing...")
    selectedOrderID = copyRange(3, 3, 3, 9, sheet)  # Change the 4 number values
    pasteOrderID = pasteRange(2, 2, 2, 8, temp_sheet, selectedOrderID)  # Change the 4 number values
    # You can save the template as another file to create a new file here too.s
    # template.save("report.xlsx")
    print("Order ID is copied and pasted!")

# def copyPhone():
# 	print("Processing...")
# 	copyPhoneNumber = copyRange(7, 3, 7, 9, sheet)
# 	pastePhoneNumber = pasteRange(3, 2, 3, 8, temp_sheet, copyPhoneNumber)
# 	print("Phone number is copied and pasted!")
#
# def copyLocation():
# 	print("Processing...")
# 	copyLocation = copyRange(8, 3, 8, 9, sheet)
# 	pasteLocation = pasteRange(4, 2, 4, 8, temp_sheet, copyLocation)
# 	print("Location is copied and pasted!")
#
# def copyAmount():
# 	print("Processing...")
# 	copyAmount = copyRange(4, 3, 4, 9, sheet)
# 	pasteAmount = pasteRange(5, 2, 5, 8, temp_sheet, copyAmount)
# 	print("Amount is copied and pasted!")

def main():
    copyOrderID()

if __name__ == '__main__':
    main()